import tempfile
from arcgis.gis import GIS
import shutil
import os
import warnings
import sys
import re
import logging, datetime
from collections import Counter
import argparse
import openpyxl as xl
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table
from openpyxl.utils import get_column_letter
from copy import copy

# Pre 3.10 surveys had a different implementation of the data validation which throws a warning message.
# This suppresses that message
warnings.simplefilter("ignore")


# Due to https://foss.heptapod.net/openpyxl/openpyxl/-/issues/1732 openpyxl will automatically set showErrorMessage and
# showInputMessage to True, this is not the case for all our data validation columns, so we need to go back and remove
# those settings for the correct data validation columns
def update_data_validation(workbook):
    ws = ["survey", "settings"]
    for worksheet in ws:
        current_ws = workbook[worksheet]
        if worksheet == "survey":
            # type column
            current_ws.data_validations.dataValidation[0].showErrorMessage = False
            current_ws.data_validations.dataValidation[0].showInputMessage = False
            # name column
            current_ws.data_validations.dataValidation[1].showInputMessage = False
            # appearance column
            current_ws.data_validations.dataValidation[6].showErrorMessage = False
            current_ws.data_validations.dataValidation[6].showInputMessage = False
            # required & read only columns
            current_ws.data_validations.dataValidation[2].showErrorMessage = False
            current_ws.data_validations.dataValidation[2].showInputMessage = False
            # bind::type column
            current_ws.data_validations.dataValidation[3].showInputMessage = False
            # bind::esri:fieldType column
            current_ws.data_validations.dataValidation[4].showInputMessage = False
            # bind::esri:fieldLength column
            current_ws.data_validations.dataValidation[5].showInputMessage = False
        elif worksheet == "settings":
            # style column
            current_ws.data_validations.dataValidation[0].showErrorMessage = False
            current_ws.data_validations.dataValidation[0].showInputMessage = False


def rebuild_table(ws2):
    try:
        # Find the table name
        ws_table = ws2.tables.items()[0][0]
        # Connect to the table
        my_table = ws2.tables[ws_table]
        # Obtain the table style info
        my_table_style = my_table.tableStyleInfo
        # Delete the table
        del ws2.tables[ws_table]

    except IndexError:
        pass

    try:
        # Define a new table given the new columns added, setting the headerRowCount to 0 disables the filters
        if ws2.max_row > 150:
            tab_rows = ws2.max_row
        else:
            tab_rows = 150
        tab = Table(displayName="{}".format(ws_table),
                    ref="$A$1:${}${}".format(get_column_letter(ws2.max_column), tab_rows), headerRowCount=0)
        # Style it the same as the old table
        tab.tableStyleInfo = my_table_style
        # Create the table
        ws2.add_table(tab)
        # print(ws2.filters)
    except:
        pass


def identify_multilingual_form(updated_col_header):
    a = False
    for x in updated_col_header:
        if x is not None:
            langs = re.search(r'\([^?!.*(.).*\1]*\)', x)
            if langs is not None:
                langs = langs[0]
                freq = Counter(langs)
                if (len(freq) == len(langs)):
                    a = True
    return a


def multilingual_rmv_def_cols(col_header, worksheet, logger):
    def_cols = ['label', 'hint', 'guidance_hint', 'required_message', 'label::language (xx)', 'hint::language (xx)',
                'guidance_hint::language (xx)', 'required_message::language (xx)']
    cois = [x for x in col_header if x in def_cols]
    for col in cois:
        col_idx = col_header.index(col) + 1
        for row in worksheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=worksheet.max_row,
                                       values_only=True):
            if len([x for x in row if x != None]) > 0:
                logger.info(
                    "A multilingual survey was detected and the {} column has values in it!\nThis column will not be deleted. Consider using a language specification for this column.".format(
                        col))
            else:
                worksheet.delete_cols(col_header.index(col) + 1, amount=1)
        col_header = [x.value for x in worksheet[1]]


def add_usr_cols(usr_cols, worksheet, wb2, ws2, header_ft, header_fill, source_workbook_header, ws1,
                 destination_workbook_header, logger):
    if worksheet == "survey":
        AB_column_dimensions = str(wb2[worksheet].column_dimensions["A"].width)  # 24
        CAE_column_dimensions = str(wb2[worksheet].column_dimensions["C"].width)  # 27
        maj_column_dimensions = str(wb2[worksheet].column_dimensions["D"].width)  # 22
        AF_column_dimensions = str(wb2[worksheet].column_dimensions["AF"].width)  # 31
        AG_column_dimensions = str(wb2[worksheet].column_dimensions["AG"].width)  # 23
        AI_column_dimensions = str(wb2[worksheet].column_dimensions["AI"].width)  # 26
    elif worksheet == "choices":
        F_column_dimensions = str(wb2[worksheet].column_dimensions["F"].width)  # 22
        maj_column_dimensions = str(wb2[worksheet].column_dimensions["A"].width)  # 18
    elif worksheet == "settings":
        maj_column_dimensions = str(wb2[worksheet].column_dimensions["A"].width)  # 18
        G_column_dimensions = str(wb2[worksheet].column_dimensions["G"].width)  # 20
        H_column_dimensions = str(wb2[worksheet].column_dimensions["H"].width)  # 29

    try:
        # Find the table name
        ws_table = ws2.tables.items()[0][0]
        # Connect to the table
        my_table = ws2.tables[ws_table]
        # Obtain the table style info
        my_table_style = my_table.tableStyleInfo
        # Delete the table
        del ws2.tables[ws_table]

    except IndexError:
        pass

    # If multilingual survey, removes the generic columns.
    if identify_multilingual_form(source_workbook_header) == True:
        multilingual_rmv_def_cols(destination_workbook_header, ws2, logger)

    for col in usr_cols:
        # Add the column in the same index
        ws2.insert_cols(col, amount=1)
        # # Give the column a name
        ws2.cell(row=1, column=col).value = usr_cols[col]
        # Style the column
        coi = ws2.cell(row=1, column=col)
        # coi.value = usr_cols[col]
        coi.font = header_ft
        coi.fill = header_fill
        coi.alignment = Alignment(horizontal="center")

        # Obtain a list of cells in the current column from the old xlsx file
        current_source_index = source_workbook_header.index(usr_cols[col])
        current_col = ws1.iter_cols(min_col=current_source_index + 1, max_col=current_source_index + 1,
                                    min_row=2, max_row=None)  # since iter_cols is a 1 based index we add 1
        # Copy the data over into the new column
        for column in current_col:
            for idx, cell in enumerate(column, 2):
                if cell.value is not None:
                    ws2.cell(row=idx, column=col).value = cell.value  # 1-indexing

    updated_col_header = [x.value for x in ws2[1]]

    if worksheet == 'survey':
        updated_dv_cols = {v: get_column_letter(k + 1) for k, v in enumerate(updated_col_header) if
                           v == 'type' or v == 'name' or v == 'required' or v == 'readonly' or v == 'bind::type' or v == 'bind::esri:fieldType' or v == 'bind::esri:fieldLength' or v == 'appearance'}
        # gives us {'type': 'A', 'name': 'B', 'appearance': 'O', 'required': 'P', 'readonly': 'U', 'bind::type': 'AE', 'bind::esri:fieldType': 'AF', 'bind::esri:fieldLength': 'AG'}
        ws2.data_validations.dataValidation[0].sqref = xl.worksheet.cell_range.MultiCellRange(
            ranges=('{}2:{}150'.format(updated_dv_cols['type'], updated_dv_cols['type'])))

        ws2.data_validations.dataValidation[1].sqref = xl.worksheet.cell_range.MultiCellRange(
            ranges=('{}2:{}150'.format(updated_dv_cols['name'], updated_dv_cols['name'])))

        ws2.data_validations.dataValidation[2].sqref = xl.worksheet.cell_range.MultiCellRange(
            ranges=('{}2:{}150'.format(updated_dv_cols['required'], updated_dv_cols['required']),
                    '{}2:{}150'.format(updated_dv_cols['readonly'], updated_dv_cols['readonly'])))

        ws2.data_validations.dataValidation[3].sqref = xl.worksheet.cell_range.MultiCellRange(
            ranges=('{}2:{}150'.format(updated_dv_cols['bind::type'], updated_dv_cols['bind::type'])))

        ws2.data_validations.dataValidation[4].sqref = xl.worksheet.cell_range.MultiCellRange(
            ranges=(
                '{}2:{}150'.format(updated_dv_cols['bind::esri:fieldType'], updated_dv_cols['bind::esri:fieldType'])))

        ws2.data_validations.dataValidation[5].sqref = xl.worksheet.cell_range.MultiCellRange(
            ranges=('{}2:{}150'.format(updated_dv_cols['bind::esri:fieldLength'],
                                       updated_dv_cols['bind::esri:fieldLength'])))

        ws2.data_validations.dataValidation[6].sqref = xl.worksheet.cell_range.MultiCellRange(
            ranges=('{}2:{}150'.format(updated_dv_cols['appearance'], updated_dv_cols['appearance'])))
    elif worksheet == 'settings':
        updated_dv_cols = {v: get_column_letter(k + 1) for k, v in enumerate(updated_col_header) if
                           v == 'style'}
        ws2.data_validations.dataValidation[0].sqref = xl.worksheet.cell_range.MultiCellRange(
            ranges=('{}2:{}150'.format(updated_dv_cols['style'], updated_dv_cols['style'])))

    try:
        # Define a new table given the new columns added, setting the headerRowCount to 0 disables the filters
        if ws2.max_row > 150:
            tab_rows = ws2.max_row
        else:
            tab_rows = 150
        tab = Table(displayName="{}".format(ws_table),
                    ref="$A$1:${}${}".format(get_column_letter(ws2.max_column), tab_rows), headerRowCount=0)
        # Style it the same as the old table
        tab.tableStyleInfo = my_table_style
        # Create the table
        ws2.add_table(tab)
        # print(ws2.filters)
    except:
        pass

    for column_cells in ws2.columns:
        col_letter = get_column_letter(column_cells[0].column)
        if (worksheet == "survey" and col_letter == "A") or (worksheet == "survey" and col_letter == "B"):
            ws2.column_dimensions[col_letter].width = AB_column_dimensions
        elif (worksheet == "survey" and col_letter == "C") or (worksheet == "survey" and col_letter == "AE"):
            ws2.column_dimensions[col_letter].width = CAE_column_dimensions
        elif worksheet == "survey" and col_letter == "AF":
            ws2.column_dimensions[col_letter].width = AF_column_dimensions
        elif worksheet == "survey" and col_letter == "AG":
            ws2.column_dimensions[col_letter].width = AG_column_dimensions
        elif worksheet == "survey" and col_letter == "AI":
            ws2.column_dimensions[col_letter].width = AI_column_dimensions
        elif worksheet == "survey":
            ws2.column_dimensions[col_letter].width = maj_column_dimensions

        elif worksheet == "choices" and col_letter == "F":
            ws2.column_dimensions[col_letter].width = F_column_dimensions
        elif worksheet == "choices":
            ws2.column_dimensions[col_letter].width = maj_column_dimensions

        elif worksheet == "settings" and col_letter == "G":
            ws2.column_dimensions[col_letter].width = G_column_dimensions
        elif worksheet == "settings" and col_letter == "H":
            ws2.column_dimensions[col_letter].width = H_column_dimensions
        elif worksheet == "settings":
            ws2.column_dimensions[col_letter].width = maj_column_dimensions


# loops through the survey, choices, and settings worksheets mapping fields from the original survey to the template,
# and copies data to those fields. The script will only be copying cells that have values
def update_survey(old_form, new_form, logger):
    # Add styling to the new fields. Source doc: https://openpyxl.readthedocs.io/en/stable/styles.html
    # header_ft = Font(bold=True, color='ffffff')
    # header_fill = PatternFill("solid", fgColor='00b050')
    # Connect to both the old XLSForm and the renamed template
    wb1 = xl.load_workbook(old_form)
    wb2 = xl.load_workbook(new_form)

    worksheets = ['survey', 'choices', 'settings']
    # Loop through the worksheets
    for worksheet in worksheets:
        # Prepare formatting based on XLSForm template
        header_ft = copy(wb2[worksheet].cell(row=1, column=1).font)
        header_fill = copy(wb2[worksheet].cell(row=1, column=1).fill)
        tab_color = copy(wb2[worksheet].sheet_properties.tabColor)
        column_dimensions = str(wb2[worksheet].column_dimensions["A"].width)

        # Connect to the worksheet in the source and destination workbooks, and identify column headers.
        ws1 = wb1[worksheet]
        source_workbook_header = [x.value for x in ws1[1]]
        ws2 = wb2[worksheet]
        destination_workbook_header = [x.value for x in ws2[1]]
        # Since the Advanced template has sample text in the survey, choices,
        # and settings worksheet we clear the text before copying content over
        ws2.delete_rows(2, 10)

        usr_cols = {k + 1: v for k, v in enumerate(source_workbook_header) if
                    v not in destination_workbook_header and v is not None and v != 'label::language1' and v != 'hint::language1' and v != 'body::esri:Parameters' and v != "allow_choice_duplicates"}

        # loop through column headers replacing min and max col with the index of the current column
        for c_name in source_workbook_header:
            if c_name in destination_workbook_header:
                # Find the index of the current column in both XLSForms
                current_source_index = source_workbook_header.index(c_name)
                current_destination_index = destination_workbook_header.index(c_name)
                # Obtain a list of cells in the current column from the old xlsx file
                current_col = ws1.iter_cols(min_col=current_source_index + 1, max_col=current_source_index + 1,
                                            min_row=2, max_row=None)  # since iter_cols is a 1 based index we add 1
                # current_col is a generator object, so we need to obtain all the cells in the object using the for
                # loops below
                for column in current_col:  # Which will always be one column
                    for idx, cell in enumerate(column, 2):
                        if cell.value is not None:
                            # Only copies cells that have values into the destination workbook
                            ws2.cell(row=idx, column=current_destination_index + 1).value = cell.value  # 1-indexing

            elif worksheet == "choices" and c_name == 'image':
                # Find the index of the current column in both XLSForms
                current_source_index = source_workbook_header.index(c_name)
                current_destination_index = destination_workbook_header.index('media::image')
                # Obtain a list of cells in the current column from the old xlsx file
                current_col = ws1.iter_cols(min_col=current_source_index + 1, max_col=current_source_index + 1,
                                            min_row=2, max_row=None)  # since iter_cols is a 1 based index we add 1
                # current_col is a generator object, so we need to obtain all the cells in the object using the for
                # loops below
                for column in current_col:  # Which will always be one column
                    for idx, cell in enumerate(column, 2):
                        if cell.value is not None:
                            # Only copies cells that have values into the destination workbook
                            ws2.cell(row=idx, column=current_destination_index + 1).value = cell.value  # 1-indexing

            elif worksheet == "choices" and c_name == 'audio':
                # Find the index of the current column in both XLSForms
                current_source_index = source_workbook_header.index(c_name)
                current_destination_index = destination_workbook_header.index('media::audio')
                # Obtain a list of cells in the current column from the old xlsx file
                current_col = ws1.iter_cols(min_col=current_source_index + 1, max_col=current_source_index + 1,
                                            min_row=2, max_row=None)  # since iter_cols is a 1 based index we add 1
                # current_col is a generator object, so we need to obtain all the cells in the object using the for
                # loops below
                for column in current_col:  # Which will always be one column
                    for idx, cell in enumerate(column, 2):
                        if cell.value is not None:
                            # Only copies cells that have values into the destination workbook
                            ws2.cell(row=idx, column=current_destination_index + 1).value = cell.value  # 1-indexing

        # Add user created columns that are not from an old template or have a null column header
        if len(usr_cols) > 0:
            add_usr_cols(usr_cols, worksheet, wb2, ws2, header_ft, header_fill, source_workbook_header, ws1,
                         destination_workbook_header, logger)

        # Extend the table if the user has more than 150 rows
        if ws2.max_row > 150:
            rebuild_table(ws2)

        # Let the user know what worksheet was just completed
        logger.info("Completed updating {} worksheet".format(worksheet))

    # Check for external_choices worksheet
    if 'external_choices' in wb1.sheetnames:
        logger.info("Copying over external choices")
        # obtain the index of the external choices worksheet, so we can recreate it in the new xlsx file
        external_idx = wb1.sheetnames.index('external_choices')
        # create the new external choices worksheet and colorize it to match the other worksheets
        external_choices = wb2.create_sheet('external_choices', external_idx)
        external_choices.sheet_properties.tabColor = tab_color
        # Make a connection to the old workbook and obtain a list of headers in the external choices worksheet
        old_x_worksheet = wb1['external_choices']
        source_x_header = [x.value for x in old_x_worksheet[1]]
        # Loop through the column headers and add them to the new xlsx file and colorize
        for idx, header in enumerate(source_x_header, 1):
            # Add the column header and colorize it
            coi = external_choices.cell(row=1, column=idx)
            coi.value = header
            coi.font = header_ft
            coi.fill = header_fill
            coi.alignment = Alignment(horizontal="center")
            # Obtain a list of cells in the current column from the old xlsx file
            current_col = old_x_worksheet.iter_cols(min_col=idx, max_col=idx, min_row=2, max_row=None)
            # Copy over cell values for each column
            for column in current_col:
                for row_idx, cell in enumerate(column, 2):
                    if cell.value is not None:
                        external_choices.cell(row=row_idx, column=idx).value = cell.value
        for column_cells in external_choices.columns:
            external_choices.column_dimensions[get_column_letter(column_cells[0].column)].width = column_dimensions
    # Update the data validation to workaround openpyxl bug
    update_data_validation(wb2)
    # Update the version worksheet
    version_ws = wb2['Version']
    coi = version_ws.cell(row=3, column=2)
    coi.value = float(coi.value)
    # Save the XLSForm
    wb2.save(new_form)
    # Close the XLSForms
    wb1.close()
    wb2.close()


def main(argv):
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsform", help="Please specify the path to the My Survey Designers folder or "
                                        "a specific survey folder")
    args = parser.parse_args()
    xlsform = args.xlsform

    # Create a temp directory in order to download the latest XLSForm template
    tmpdir = tempfile.TemporaryDirectory()
    # Extract the directory path
    xls_directory = tmpdir.name
    # Make a GIS Connection, since the Advanced XLSForm template is public we don't need credentials to download it
    gis = GIS()
    advanced_template = gis.content.get("3db671560a72427b8d778a2da0979773")
    # Download the Advanced XLSForm template and use the item name as the name of the file downloaded
    advanced_template_name = advanced_template.name
    advanced_template_directory = advanced_template.download(xls_directory, advanced_template_name)

    timestamp = datetime.datetime.now()

    head, tail = os.path.split(xlsform)

    debug_path = os.path.join(head, "debug")
    if not os.path.exists(debug_path):
        os.makedirs(debug_path)
    update_path = os.path.join(debug_path, "template_updater")
    if not os.path.exists(update_path):
        os.makedirs(update_path)

    # Set up logging
    logging.basicConfig(level=logging.INFO)
    logger = logging.getLogger(__name__)
    logFileName = "template_update_{}".format(timestamp.strftime('%Y-%m-%d %H-%M-%S'))
    fileHandler = logging.handlers.RotatingFileHandler('{}/{}.log'.format(update_path, logFileName), maxBytes=100000,
                                                       backupCount=5)
    formatter = logging.Formatter(
        '%(asctime)s %(levelname)s %(relativeCreated)d \n%(filename)s %(module)s %(funcName)s %(lineno)d \n%(message)s\n')
    fileHandler.setFormatter(formatter)
    logger.addHandler(fileHandler)
    logger.info('Script starting at {}'.format(str(datetime.datetime.now())))

    # Obtain the name of the XLSForm
    xlsx_name = os.path.splitext(tail)[0]
    # Rename the XLSForm to <name>_old so the user has a backup of their XLSForm
    shutil.move(os.path.join(head, xlsx_name + ".xlsx"), update_path)
    old_xlsx = os.path.join(update_path,
                            xlsx_name + "_{}.xlsx".format(timestamp.strftime('%Y_%m_%d_%H_%M_%S')))
    os.rename(os.path.join(update_path, xlsx_name + ".xlsx"), old_xlsx)
    logger.info("Made a backup of the XLSForm in:\n{}".format(os.path.join(update_path, xlsx_name + "_{}.xlsx".format(
        timestamp.strftime('%Y_%m_%d_%H_%M_%S')))))
    # Copy the blank XLSForm template and rename it the same as the survey
    target_xlsx = shutil.copy(advanced_template_directory, xlsform)
    # Let the user know what survey is currently being worked on
    logger.info("Currently working on survey: {}".format(xlsx_name))
    # We then update the form copying everything from the _old form.
    try:
        update_survey(old_xlsx, target_xlsx, logger)
    except:
        logger.info("Unable to process survey: {}".format(xlsx_name))


if __name__ == "__main__":
    main(sys.argv[1:])
